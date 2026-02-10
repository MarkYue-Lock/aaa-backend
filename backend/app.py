import os
import random
import pandas as pd
from flask import Flask, request, jsonify, Response, stream_with_context
from flask_cors import CORS
import requests
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import tempfile

app = Flask(__name__)

# =========================================================
# 1. CONFIGURATION & CORS
# =========================================================
# Allow your GitHub Pages to talk to this backend
CORS(app, resources={r"/*": {"origins": "*"}})

# Dify Configuration (For the Retail Chatbot side)
DIFY_API_KEY = os.environ.get("DIFY_API_KEY", "YOUR_DIFY_API_KEY_HERE")
DIFY_BASE_URL = "https://api.dify.ai/v1"

# =========================================================
# 2. YOUR HOMEPORT QUALIFIER CLASS
# =========================================================
class HomeportQualifier:
    ASSET_COEFFICIENTS = [
        ("Checking/Saving/CD", 1.00),
        ("Stocks, Bonds, Mutual funds", 0.90),
        ("Retirement >= 59 1/2", 0.90),
        ("Retirement < 59 1/2", 0.70),
        ("Life Insurance(cash surrender value-loans)", 0.90),
        ("529 account(Soly owned)", 0.60),
    ]
    
    DEFAULT_CELL_MAP = {
        "Sub_Add": "E6",
        "Loan_Purpose": "E7",
        "Sub_App_Value": "E12",
        "Sub_Pur_Value": "E13",
        "DP_CC": "E15",
    }

    def __init__(self, excel_path, sheet_name="Version#1", threshold=2800):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.base_threshold = threshold
        
        self.wb = None
        self.data = {}
        self.results = {}

    def load_workbook(self):
        if not self.wb:
            try:
                self.wb = load_workbook(self.excel_path, data_only=True)
            except FileNotFoundError:
                raise FileNotFoundError(f"cant find: {self.excel_path}")
        return self.wb

    @staticmethod
    def _excel_mapping_to_df(wb, sheet_name, cell_mapping):
        ws = wb[sheet_name]
        result = {}
        for base_name, location in cell_mapping.items():
            if ":" in location:
                min_col, min_row, max_col, max_row = range_boundaries(location)
                idx = 1
                for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    for cell in row:
                        col_name = f"{base_name}{idx}"
                        result[col_name] = cell.value
                        idx += 1
            else:
                result[base_name] = ws[location].value
        return pd.DataFrame([result])

    @staticmethod
    def _read_excel_range_with_header(wb, sheet_name, cell_range):
        ws = wb[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)

        data = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            data.append(list(row))

        if not data:
            return pd.DataFrame()

        raw_headers = data[0]
        headers = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(raw_headers)]

        df = pd.DataFrame(data[1:], columns=headers)
        
        df = df.dropna(how="all", axis=1).dropna(how="all", axis=0)
        return df

    def _process_basic_info(self):
        wb = self.load_workbook()
        df = self._excel_mapping_to_df(wb, self.sheet_name, self.DEFAULT_CELL_MAP)
        
        loan_purpose = str(df['Loan_Purpose'].iloc[0]).strip() if pd.notna(df['Loan_Purpose'].iloc[0]) else ""
        
        df["Sub_Value"] = (df[["Sub_App_Value", "Sub_Pur_Value"]].apply(pd.to_numeric, errors="coerce").min(axis=1))
        raw_sub_val = df['Sub_Value'].iloc[0] if pd.notna(df['Sub_Value'].iloc[0]) else 0
        
        self.data['sub_val'] = 0 if loan_purpose == "Purchase" else raw_sub_val
        self.data['sub_add'] = str(df["Sub_Add"].iloc[0]).strip().lower()
        self.data['dp_cc'] = pd.to_numeric(df['DP_CC'].iloc[0], errors='coerce') if pd.notna(df['DP_CC'].iloc[0]) else 0

    def _process_assets(self):
        wb = self.load_workbook()
        asset_df = self._read_excel_range_with_header(wb, self.sheet_name, "B32:G37")
        
        if asset_df.empty:
            self.data['liq_asset'] = 0
            return

        cols_to_numeric = ['Balance', "Borrowers' ownership of account (%)"]
        for col in cols_to_numeric:
            if col in asset_df.columns:
                asset_df[col] = pd.to_numeric(asset_df[col], errors='coerce').fillna(0)
            else:
                asset_df[col] = 0

        cf_df = pd.DataFrame(self.ASSET_COEFFICIENTS, columns=["Type", "CF"])
        
        asset_df['Type'] = asset_df['Type'].astype(str).str.strip()
        
        asset_merged = asset_df.merge(cf_df, on="Type", how="left")
        
        asset_merged['CF'] = asset_merged['CF'].fillna(0)

        asset_merged['bal_adj'] = (
            asset_merged['Balance'] * asset_merged["Borrowers' ownership of account (%)"] * asset_merged['CF']
        )/100
        self.data['liq_asset'] = asset_merged['bal_adj'].sum()

    def _process_reo(self):
        wb = self.load_workbook()
        reo_df = self._read_excel_range_with_header(wb, self.sheet_name, "B47:K55")
        
        if reo_df.empty:
            self.data['non_sub_val'] = 0
            self.data['non_sub_val_adj'] = 0
            self.data['sub_pitia'] = 0
            self.data['non_sub_pitiasm'] = 0
            return

        required_cols = ["Address", "Zillow Value", "Ownership(%)"]
        sub_add = self.data.get('sub_add', "")
        reo_df["Address"] = reo_df["Address"].astype(str)
        reo_df["sub_prpty"] = reo_df["Address"].str.contains(sub_add, case=False, na=False, regex=False).map({True: "Y", False: "N"})
        
        non_sub_mask = reo_df["sub_prpty"] == "N"
        
        numeric_cols = ["Zillow Value", "Ownership(%)", "Monthly PI", "Monthly Tax", "Monthly Ins", "Monthly HOA", "Monthly Solar", "Monthly MI"]
        
        col_map = {}
        for col in reo_df.columns:
            clean_name = " ".join(col.split())
            col_map[col] = clean_name
        
        reo_df.rename(columns=col_map, inplace=True)
        
        for col in numeric_cols:
            if col in reo_df.columns:
                reo_df[col] = pd.to_numeric(reo_df[col], errors='coerce').fillna(0)
            else:
                reo_df[col] = 0

        non_sub_val = (
            reo_df.loc[non_sub_mask, "Zillow Value"] * reo_df.loc[non_sub_mask, "Ownership(%)"]
        ).sum() / 100
        self.data['non_sub_val'] = non_sub_val
        self.data['non_sub_val_adj'] = non_sub_val * 0.9

        # 5. PITIA
        pitia_cols_sub = ["Monthly PI", "Monthly Tax", "Monthly Ins", "Monthly HOA"]
        self.data['sub_pitia'] = reo_df.loc[reo_df["sub_prpty"] == "Y", pitia_cols_sub].sum().sum()

        pitia_cols_non_sub = ["Monthly PI", "Monthly Tax", "Monthly Ins", "Monthly HOA", "Monthly Solar", "Monthly MI"]
        self.data['non_sub_pitiasm'] = reo_df.loc[non_sub_mask, pitia_cols_non_sub].sum().sum()

    def _process_others(self):
        wb = self.load_workbook()
        
        # Gift
        gift_df = self._read_excel_range_with_header(wb, self.sheet_name, "B39:G43")
        if not gift_df.empty and 'Amount' in gift_df.columns:
            self.data['gift_amt'] = pd.to_numeric(gift_df['Amount'], errors='coerce').fillna(0).sum()
        else:
            self.data['gift_amt'] = 0
        
        # Other Debt
        debt_df = self._read_excel_range_with_header(wb, self.sheet_name, "B21:G29")
        if not debt_df.empty and 'Monthly Payment' in debt_df.columns:
            self.data['mthly_pmt'] = pd.to_numeric(debt_df['Monthly Payment'], errors='coerce').fillna(0).sum()
        else:
            self.data['mthly_pmt'] = 0

    def calculate(self):
        self._process_basic_info()
        self._process_assets()
        self._process_reo()
        self._process_others()
        
        numerator = (
            self.data['liq_asset'] + 
            self.data['sub_val'] + 
            self.data['non_sub_val_adj'] - 
            self.data['dp_cc'] + 
            self.data['gift_amt']
        )
        self.results['mthly_income'] = numerator / 36
        
        self.results['mthly_debt'] = (
            self.data['mthly_pmt'] + 
            self.data['sub_pitia'] + 
            self.data['non_sub_pitiasm']
        )
        
        self.results['residual'] = self.results['mthly_income'] - self.results['mthly_debt']

        random_premium = round(random.uniform(0, 100), 2)
        self.results['final_threshold'] = self.base_threshold + random_premium
        self.results['random_premium'] = random_premium
        
        return self.results

    def evaluate(self):
        if not self.results:
            self.calculate()
            
        res = self.results
        data = self.data
        lines = []
        
        # --- CONFIGURATION ---
        WIDTH = 50 
        SEP = "=" * WIDTH
        DASH = "-" * WIDTH
        
        # --- 1. MAIN HEADER ---
        lines.append("")
        lines.append(SEP)
        header_content = f"{'>>> HOMEPORT QUALIFICATION REPORT <<<':^{WIDTH}}"
        lines.append(f"<strong>{header_content}</strong>")
        lines.append(SEP)

        # --- 2. FINANCIAL DETAILS ---
        row = lambda k, v: f"{k:<30} {v:>18,.2f}"

        lines.append(row("Liquid Assets:", data['liq_asset']))
        lines.append(row("Subject Value Used:", data['sub_val']))
        lines.append(row("Non-Sub Value (90%):", data.get('non_sub_val_adj', 0))) 
        lines.append(row("DP & Closing Cost:", data['dp_cc']))
        lines.append(row("Gift Amount:", data['gift_amt']))
        
        lines.append(DASH)
        lines.append(row("Monthly Income:", res['mthly_income']))
        lines.append(DASH)
        
        lines.append(row("Payment (except REO):", data['mthly_pmt']))
        lines.append(row("Sub PITIA:", data['sub_pitia']))
        lines.append(row("Non-Sub PITIASM:", data.get('non_sub_pitiasm', 0)))

        lines.append(DASH)
        lines.append(row("Monthly Debt:", res['mthly_debt']))
        lines.append(SEP)
        
        lines.append(row("Residual Income:", res['residual']))
        lines.append(f"{'Base Threshold:':<30} {self.base_threshold:>18}") 
        lines.append(f"{'Random Premium:':<30} {res['random_premium']:>18}")
        lines.append(row("Final Threshold:", res['final_threshold']))
        lines.append(SEP)

        # --- 3. RESULT LOGIC ---
        status_content = f"{'QUALIFIER STATUS':^{WIDTH}}"
        lines.append(f"<strong>{status_content}</strong>")
        lines.append(DASH)
        
        if res['final_threshold'] > res['residual']:
            # FAIL STATE
            fail_title = f"{'xx NOT ELIGIBLE xx':^{WIDTH}}"
            lines.append(f"<strong>{fail_title}</strong>")
            lines.append("")
            lines.append("You <strong>MAY NOT BE ELIGIBLE</strong> for Homeport Program.")
            
        elif data.get('gift_amt', 0) > data.get('dp_cc', 0):
            # WARNING STATE
            warn_title = f"{'!! ELIGIBILITY WARNING !!':^{WIDTH}}"
            lines.append(f"<strong>{warn_title}</strong>")
            lines.append("")
            lines.append("<strong>Gift Amount Exceeds Down Payment/Closing Costs.</strong>")
            
        else:
            # SUCCESS STATE
            success_title = f"{'>> POTENTIALLY ELIGIBLE <<':^{WIDTH}}"
            lines.append(f"<strong>{success_title}</strong>")
            lines.append("")
            lines.append("You <strong>MAY BE ELIGIBLE</strong> for Homeport Program.")

        lines.append(SEP)

        # --- 4. DISCLAIMER ---
        lines.append("This is a preliminary calculation — this <strong>does not</strong>")
        lines.append("<strong>constitute loan approval</strong>. Results are subject to")
        lines.append("change during underwriting. New/updated")
        lines.append("documentation or revised loan details may require")
        lines.append("adjusted qualifying calculations, all subject to")
        lines.append("underwriting review. Any loan information not on")
        lines.append("the Submission Ticket must adhere to the AAA")
        lines.append("Matrix and rate sheet.")
        
        lines.append(SEP)

        return "\n".join(lines)

# =========================================================
# 3. ROUTES
# =========================================================

@app.route('/', methods=['GET'])
def health_check():
    return "Backend is active and running!", 200

# Route A: HomePort Analysis
@app.route('/api/homeport/analyze', methods=['POST'])
def analyze_homeport():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    file = request.files['file']
    
    # Save uploaded file to a temporary location for processing
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name

    try:
        # Run your logic
        hq = HomeportQualifier(tmp_path)
        report = hq.evaluate()
        
        # Cleanup
        os.remove(tmp_path)
        
        return report, 200
        
    except Exception as e:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        return jsonify({"error": str(e)}), 500

# Route B: Dify Chat (Retail Channel)
@app.route('/api/chat-stream', methods=['POST'])
def chat_stream():
    try:
        data = request.json
        headers = {
            'Authorization': f'Bearer {DIFY_API_KEY}',
            'Content-Type': 'application/json'
        }
        
        resp = requests.post(
            f"{DIFY_BASE_URL}/chat-messages",
            json=data,
            headers=headers,
            stream=True
        )

        def generate():
            for chunk in resp.iter_content(chunk_size=1024):
                if chunk: yield chunk

        return Response(stream_with_context(generate()), content_type=resp.headers.get('Content-Type'))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Route C: Dify File Upload
@app.route('/api/files/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    
    file = request.files['file']
    user = request.form.get('user', 'default-user')

    try:
        headers = {'Authorization': f'Bearer {DIFY_API_KEY}'}
        files = {'file': (file.filename, file.stream, file.content_type)}
        data = {'user': user}
        
        resp = requests.post(
            f"{DIFY_BASE_URL}/files/upload",
            headers=headers,
            files=files,
            data=data
        )
        return jsonify(resp.json()), resp.status_code
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5001))
    app.run(host='0.0.0.0', port=port)